import os
import time
import math
import re
import uuid
import shutil
import mimetypes
import tempfile
from urllib.parse import urlparse, parse_qs

import requests
import pandas as pd
from celery import shared_task, group
from celery.utils.log import get_task_logger
from django.conf import settings
from django.utils import timezone
from django.core.mail import EmailMultiAlternatives, get_connection
from django.template.loader import render_to_string
from django.utils.html import strip_tags, escape

from .models import EmailFile, EmailRecord, SMTPAccount

logger = get_task_logger(__name__)

# -----------------------------
# Attachment download defaults (self-contained)
# -----------------------------
# tmp root: <BASE_DIR>/email_attachments_tmp or /tmp/email_attachments_tmp
ATT_TMP_ROOT = os.path.join(getattr(settings, "BASE_DIR", "/tmp"), "email_attachments_tmp")
ATT_TIMEOUT = 30  # seconds per request
ATT_PER_FILE_MAX_MB = 20
ATT_TOTAL_MAX_MB = 25
ATT_ALLOWED_SCHEMES = {"http", "https"}

# bytes
ATT_PER_FILE_MAX = ATT_PER_FILE_MAX_MB * 1024 * 1024
ATT_TOTAL_MAX = ATT_TOTAL_MAX_MB * 1024 * 1024

os.makedirs(ATT_TMP_ROOT, exist_ok=True)


def _normalize_attachments(value) -> str:
    """
    Accepts a cell value from the 'Attachments' column and returns a
    comma-separated string of validated http(s) URLs.

    - Splits on commas (also tolerates newlines/semicolons by converting to commas)
    - Trims whitespace
    - Filters out empty items and non-http(s) schemes
    - Handles None/NaN gracefully
    """
    if value is None:
        return ""

    # Handle pandas NaN (float('nan')) and string "nan"
    if isinstance(value, float) and math.isnan(value):
        return ""
    s = str(value).strip()
    if s.lower() == "nan":
        return ""

    # Be tolerant to different separators
    s = s.replace("\n", ",").replace(";", ",")
    parts = [p.strip() for p in s.split(",")]

    out = []
    for u in parts:
        if not u:
            continue
        try:
            p = urlparse(u)
            if p.scheme in ATT_ALLOWED_SCHEMES:
                out.append(u)
        except Exception:
            # Ignore malformed URLs
            continue
    return ",".join(out)


# -----------------------------
# Google Drive helpers
# -----------------------------
_DRIVE_FILE_RE = re.compile(r"/file/d/([A-Za-z0-9_-]+)/")


def _drive_direct_url(url: str) -> str:
    """
    Convert common Google Drive share links into direct-download links.
    - https://drive.google.com/file/d/<ID>/view -> https://drive.google.com/uc?export=download&id=<ID>
    - https://drive.google.com/open?id=<ID>     -> same
    Leaves other URLs untouched.
    """
    try:
        parsed = urlparse(url)
        if "drive.google.com" not in parsed.netloc:
            return url

        # Already a uc link
        if parsed.path.startswith("/uc"):
            return url

        # /file/d/<ID>/view
        m = _DRIVE_FILE_RE.search(parsed.path)
        if m:
            file_id = m.group(1)
            return f"https://drive.google.com/uc?export=download&id={file_id}"

        # /open?id=<ID>
        qs = parse_qs(parsed.query or "")
        if "id" in qs and qs["id"]:
            file_id = qs["id"][0]
            return f"https://drive.google.com/uc?export=download&id={file_id}"
    except Exception:
        return url
    return url


def _safe_filename_from_url(url: str) -> str:
    """Derive a safe filename from the URL path, prefixed with a short UUID."""
    path = urlparse(url).path
    base = os.path.basename(path) or "attachment"
    base = base.split("?")[0].split("#")[0]
    base = base.replace("/", "_").replace("\\", "_")
    return f"{uuid.uuid4().hex[:8]}_{base}"


def _download_to_temp(url: str, tmp_dir: str):
    """
    Download a single URL to tmp_dir with size/time limits.
    Returns dict: {path, filename, mimetype, size}. Raises on failure/limits.
    """
    u = _drive_direct_url(url)
    headers = {"User-Agent": "MailAttachmentFetcher/1.0"}
    with requests.get(u, stream=True, timeout=ATT_TIMEOUT, headers=headers) as r:
        r.raise_for_status()

        # Determine filename (Content-Disposition > URL path)
        filename = _safe_filename_from_url(u)
        cd = r.headers.get("Content-Disposition", "")
        m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', cd)
        if m:
            candidate = m.group(1).strip(' "\'').replace("/", "_").replace("\\", "_")
            if candidate:
                filename = f"{uuid.uuid4().hex[:8]}_{candidate}"

        # Determine mimetype
        mimetype = r.headers.get("Content-Type")
        if not mimetype or ";" in str(mimetype):
            mimetype = (mimetypes.guess_type(filename)[0]) or "application/octet-stream"

        # Enforce Content-Length if present
        total = 0
        if r.headers.get("Content-Length"):
            try:
                cl = int(r.headers["Content-Length"])
                if cl > ATT_PER_FILE_MAX:
                    raise ValueError(f"Attachment too large (>{ATT_PER_FILE_MAX_MB}MB): {url}")
            except Exception:
                pass

        # Stream to disk with rolling size check
        os.makedirs(tmp_dir, exist_ok=True)
        out_path = os.path.join(tmp_dir, filename)
        with open(out_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if not chunk:
                    continue
                total += len(chunk)
                if total > ATT_PER_FILE_MAX:
                    raise ValueError(f"Attachment exceeded per-file limit (>{ATT_PER_FILE_MAX_MB}MB): {url}")
                f.write(chunk)

    return {"path": out_path, "filename": filename, "mimetype": mimetype, "size": total}


def _prepare_attachments_temp(urls) -> tuple[list, list, str]:
    """
    Given a list of URLs, download each into a per-email temp directory.
    Returns (attachments_meta, errors, temp_dir)
      - attachments_meta: list of dicts as returned by _download_to_temp
      - errors: list of error strings for any failures
      - temp_dir: the created temporary directory to clean up later
    """
    if not urls:
        return [], [], None

    temp_dir = tempfile.mkdtemp(prefix="mail_", dir=ATT_TMP_ROOT)
    errors = []
    metas = []
    total_size = 0

    for url in urls:
        try:
            meta = _download_to_temp(url, temp_dir)
            total_size += meta["size"]
            if total_size > ATT_TOTAL_MAX:
                errors.append(f"Total attachments exceeded limit (>{ATT_TOTAL_MAX_MB}MB). Skipped remaining.")
                break
            metas.append(meta)
        except Exception as e:
            errors.append(f"{url}: {str(e)}")

    return metas, errors, temp_dir


# -----------------------------
# HTML detection for body rendering
# -----------------------------
_HTML_TAG_RE = re.compile(r'</?[a-z][\s\S]*?>', re.IGNORECASE)


def _looks_like_html(text: str) -> bool:
    """Heuristic: does the string contain any HTML tag-like patterns?"""
    if not text:
        return False
    return bool(_HTML_TAG_RE.search(text))


# -----------------------------
# XLSX rich-text helpers (best-effort)
# -----------------------------
def _xlsx_cell_to_string_or_html(cell) -> str:
    """
    For .xlsx bodies:
      - If the cell contains explicit HTML (e.g., pasted HTML), return it unchanged.
      - If the cell has rich text runs (bold/italic/underline), convert to simple HTML.
      - Otherwise, return plain text (with \u2028 normalized to '\n').
    """
    v = cell.value
    if v is None:
        return ""

    s = str(v)
    if _looks_like_html(s):
        # Already HTML-ish, keep as-is
        return s

    # Try to handle openpyxl rich text (if available) → to HTML
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock  # type: ignore
        if isinstance(v, CellRichText):
            parts = []
            for block in v:  # iterate TextBlock runs
                txt = escape(getattr(block, "text", "") or "")
                if getattr(block, "bold", False):
                    txt = f"<b>{txt}</b>"
                if getattr(block, "italic", False):
                    txt = f"<i>{txt}</i>"
                if getattr(block, "underline", False):
                    txt = f"<u>{txt}</u>"
                parts.append(txt)
            return "".join(parts)
    except Exception:
        # If reading rich text not supported, fall back to plain
        pass

    # Plain text fallback: normalize special line separators, keep raw newlines
    return s.replace("\u2028", "\n")


def _read_defaults_and_rows(file_path: str):
    """
    Unified reader:
      - For .xlsx: use openpyxl to best-effort preserve Body formatting.
      - For .csv/.xls: use pandas; rich formatting is not present in these formats,
        but HTML typed into cells will be preserved as text.
    Returns (rows: list[dict], default_subject: str, default_body: str)
    where each dict minimally has keys: Name, Email, Subject, Body, Attachments
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook  # lazy import
        except Exception as e:
            logger.warning(f"openpyxl not available, falling back to pandas for {file_path}: {e}")
            ext = ".xlsfallback"  # force pandas path

    if ext == ".xlsx":
        wb = load_workbook(file_path, data_only=True)  # read-only False to access rich text
        ws = wb.active

        # Build header map (strip spaces)
        header_map = {}
        for col in range(1, ws.max_column + 1):
            key = ws.cell(row=1, column=col).value
            if key is not None:
                header_map[str(key).strip()] = col

        def col(name):
            return header_map.get(name)

        # Defaults from first data row (row 2)
        default_subject = ""
        default_body = ""
        if col("Subject"):
            default_subject = str(ws.cell(row=2, column=col("Subject")).value or "").strip()
        if col("Body"):
            default_body = _xlsx_cell_to_string_or_html(ws.cell(row=2, column=col("Body")))

        rows = []
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(row=r, column=col("Name")).value or "").strip() if col("Name") else ""
            email = str(ws.cell(row=r, column=col("Email")).value or "").strip() if col("Email") else ""
            subject = str(ws.cell(row=r, column=col("Subject")).value or "").strip() if col("Subject") else ""
            # We will use default_body later; still read per-row body in case you want to switch strategy
            body_cell = ws.cell(row=r, column=col("Body")) if col("Body") else None
            body = _xlsx_cell_to_string_or_html(body_cell) if body_cell else ""
            attachments_raw = ws.cell(row=r, column=col("Attachments")).value if col("Attachments") else None
            rows.append({
                "Name": name,
                "Email": email,
                "Subject": subject,
                "Body": body,
                "Attachments": attachments_raw,
            })
        wb.close()
        return rows, default_subject, default_body

    # Fallback: CSV / XLS (pandas)
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path)
    elif file_path.endswith(('.xls', '.xlsx')):  # .xlsx here only if openpyxl import failed
        df = pd.read_excel(file_path)
    else:
        raise ValueError("Unsupported file format. Only CSV, XLS, and XLSX are allowed.")

    if df.empty:
        return [], "", ""

    try:
        df.columns = df.columns.str.strip()
    except Exception:
        pass

    first_row = df.iloc[0]
    default_subject = str(first_row.get('Subject', '')).strip()
    # Normalize special unicode line separator to newline; preserve any HTML typed by user
    default_body = str(first_row.get('Body', '')).replace('\u2028', '\n')

    rows = []
    for _, row in df.iterrows():
        rows.append({
            "Name": str(row.get('Name', '')).strip(),
            "Email": str(row.get('Email', '')).strip(),
            "Subject": str(row.get('Subject', '')).strip(),
            "Body": str(row.get('Body', '')).replace('\u2028', '\n'),
            "Attachments": row.get('Attachments', None),
        })
    return rows, default_subject, default_body


@shared_task(bind=True)
def process_uploaded_file(self, email_file_id):
    logger.info(f"[TASK STARTED] Processing file: {email_file_id}")
    try:
        email_file = EmailFile.objects.get(id=email_file_id)
        file_path = email_file.file.path

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Uploaded file not found at path: {file_path}")

        # Read rows + defaults with best-effort formatting preservation
        rows, default_subject, default_body = _read_defaults_and_rows(file_path)

        if not rows:
            logger.warning(f"Uploaded file {file_path} is empty.")
            return f"No rows to process for file ID {email_file_id}."

        created_count = 0
        for row in rows:
            email = row.get("Email", "").strip()
            if not email:
                continue

            attachments_urls = _normalize_attachments(row.get("Attachments", None))

            # Keep existing behavior: use subject/body from the first data row as defaults
            EmailRecord.objects.create(
                file=email_file,
                name=row.get("Name", "").strip(),
                email=email,
                subject=default_subject,
                body=default_body,  # can be HTML (xlsx rich text → HTML) or plain (csv/xls)
                cc='',
                bcc='',
                attachments_urls=attachments_urls,
                is_sent=False
            )
            created_count += 1

        logger.info(f"[TASK COMPLETED] Created {created_count} records for file ID {email_file_id}")
        return f"Processed file ID {email_file_id} with {created_count} records."

    except EmailFile.DoesNotExist:
        logger.error(f"EmailFile with ID {email_file_id} does not exist.")
        return f"EmailFile with ID {email_file_id} not found."

    except Exception as e:
        logger.exception(f"Exception during file processing: {str(e)}")
        return f"Error processing file ID {email_file_id}: {str(e)}"


@shared_task(bind=True, max_retries=3)
def send_emails_for_file(self, email_file_id):
    """
    Parallelize sending by fanning out one Celery task per record.
    Leverages worker concurrency instead of serial loop.
    """
    try:
        file = EmailFile.objects.get(id=email_file_id)
        smtp = SMTPAccount.objects.last()

        # Reset daily counters if day changed (coarse-grained; per-record task also checks)
        if smtp and smtp.last_reset.date() < timezone.now().date():
            smtp.emails_sent_today = 0
            smtp.rate_limited = False
            smtp.last_reset = timezone.now()
            smtp.save()

        # Get unsent records
        record_ids = list(file.email_records.filter(is_sent=False).values_list('id', flat=True))
        if not record_ids:
            return f"No pending emails for file ID {file.id}"

        # Fan-out to parallel tasks (respecting your Celery worker pool size)
        job = group(send_email_record.s(rid) for rid in record_ids)
        job.apply_async()

        return f"Queued {len(record_ids)} emails for file ID {file.id}"

    except Exception as e:
        logger.exception(f"Fatal error queueing emails: {str(e)}")
        return f"Fatal error sending emails: {str(e)}"


@shared_task(bind=True)
def send_email_record(self, record_id):
    """
    Sends a single email (HTML + plain alternative), downloads/attaches files,
    and cleans up temporary files. Runs safely in parallel across workers.
    """
    try:
        record = EmailRecord.objects.get(id=record_id)

        if record.is_sent:
            return f"Email {record.email} already sent."

        smtp = SMTPAccount.objects.last()
        if not smtp:
            return "No SMTP account configured."

        # Simple daily reset (race tolerant; worst-case double reset within day boundary)
        if smtp.last_reset.date() < timezone.now().date():
            smtp.emails_sent_today = 0
            smtp.rate_limited = False
            smtp.last_reset = timezone.now()
            smtp.save()

        if smtp.rate_limited or smtp.emails_sent_today >= 500:
            smtp.rate_limited = True
            smtp.save()
            return f"Gmail quota reached for {record.file.user.email}"

        connection = None
        temp_dir = None
        try:
            connection = get_connection(
                host=smtp.email_host,
                port=smtp.email_port,
                username=smtp.email_host_user,
                password=smtp.email_host_password,
                use_tls=smtp.use_tls
            )

            subject = record.subject or "No Subject"
            raw_body = record.body or ""
            context = {
                'name': record.name,
                'body': raw_body,
                'subject': subject,
                'is_html': _looks_like_html(raw_body),
            }

            # Render template once; send as multipart/alternative
            html_content = render_to_string("emails/default_email.html", context)
            plain_content = strip_tags(html_content)

            msg = EmailMultiAlternatives(
                subject=subject,
                body=plain_content,  # plain text part
                from_email=smtp.email_host_user,
                to=[record.email],
                cc=record.cc.split(',') if record.cc else [],
                bcc=record.bcc.split(',') if record.bcc else [],
                connection=connection
            )
            msg.attach_alternative(html_content, "text/html")

            # Download + attach files (per-email temp dir)
            attachments_meta, dl_errors, temp_dir = _prepare_attachments_temp(record.attachments_list)
            for meta in attachments_meta:
                with open(meta["path"], "rb") as fh:
                    msg.attach(meta["filename"], fh.read(), meta["mimetype"])

            msg.send()

            # Mark sent + notes
            record.is_sent = True
            record.send_attempts += 1
            record.last_sent_at = timezone.now()
            if dl_errors:
                note = " | ".join(dl_errors)[:500]
                record.error_message = (record.error_message or "")
                if record.error_message:
                    record.error_message += " | "
                record.error_message += f"Attachment notes: {note}"
            else:
                record.error_message = ''
            record.save()

            # Update quota (coarse; in parallel you may see slight races which are acceptable for soft caps)
            smtp.emails_sent_today += 1
            smtp.save()

            return f"Email sent to {record.email}"

        except Exception as e:
            error_msg = str(e)[:500]
            logger.error(f"Error sending single email: {error_msg}")
            record.send_attempts += 1
            record.last_sent_at = timezone.now()
            # bubble up attachment errors if any
            if 'dl_errors' in locals() and dl_errors:
                error_msg = f"Attachment errors: {' | '.join(dl_errors)} | Send error: {error_msg}"
            record.error_message = error_msg
            record.save()
            return f"Error sending email to {record.email}: {error_msg}"
        finally:
            if temp_dir and os.path.isdir(temp_dir):
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                except Exception as ce:
                    logger.warning(f"Failed to cleanup temp dir {temp_dir}: {ce}")
            if connection:
                try:
                    connection.close()
                except Exception:
                    pass

    except Exception as e:
        error_msg = str(e)[:500]
        logger.error(f"Fatal error sending email: {error_msg}")
        return f"Fatal error sending email: {error_msg}"
